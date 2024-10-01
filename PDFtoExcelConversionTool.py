import os
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

def pdf_to_excel(pdf_path, excel_path):
    """PDFファイルをExcelファイルに変換する関数"""
    try:
        with pdfplumber.open(pdf_path) as pdf:  # PDFを開く
            wb = Workbook()  # 新しいExcelワークブックを作成
            ws = wb.active  # アクティブなシートを取得

            # PDFの各ページをループ処理
            for page_number, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()  # ページからテキストを抽出
                if text:  # テキストが存在する場合
                    for line in text.split('\n'):  # テキストを行に分ける
                        ws.append([line])  # 行をExcelシートに追加

                    ws.append([f'--- Page {page_number} ---'])  # ページの区切りを追加

                # 各行のフォントスタイルを設定
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for cell in row:
                        cell.font = Font(size=12, color='000000')  # フォントサイズと色を設定

            wb.save(excel_path)  # Excelファイルを指定したパスに保存
    except Exception as e:
        # エラーメッセージを表示し、処理を続行
        messagebox.showerror("エラー", f"PDFをExcelに変換中にエラーが発生しました:\n{str(e)}")

def convert_selected_pdf(pdf_path):
    """選択されたPDFファイルをExcelに変換する関数"""
    try:
        # 保存先とファイル名を指定するダイアログを表示
        excel_path = filedialog.asksaveasfilename(
            title="保存先とファイル名を指定",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=os.path.basename(pdf_path).replace(".pdf", ".xlsx")  # 初期ファイル名をPDF名に設定
        )
        if excel_path:  # ユーザーがファイル名を指定した場合
            pdf_to_excel(pdf_path, excel_path)  # PDFをExcelに変換
            # 完了メッセージを表示
            messagebox.showinfo("完了", f"{os.path.basename(pdf_path)} がExcelに変換されました。")
    except Exception as e:
        # エラーメッセージを表示し、処理を続行
        messagebox.showerror("エラー", f"ファイルの保存中にエラーが発生しました:\n{str(e)}")

def select_pdf_file():
    """PDFファイルを選択するためのダイアログを表示する関数"""
    file_path = filedialog.askopenfilename(title="変換したいPDFファイルを選択", filetypes=[("PDF Files", "*.pdf")])
    return file_path  # 選択されたPDFファイルのパスを返す

# メイン処理
def main():
    """プログラムのメイン処理"""
    root = tk.Tk()
    root.withdraw()  # GUIウィンドウを表示しない

    try:
        pdf_file = select_pdf_file()  # PDFファイルの選択
        if not pdf_file:  # ファイルが選択されていない場合
            messagebox.showerror("エラー", "PDFファイルが選択されていません。")  # エラーメッセージを表示
            return

        convert_selected_pdf(pdf_file)  # 選択されたPDFファイルをExcelに変換
    except Exception as e:
        # エラーメッセージを表示し、処理を続行
        messagebox.showerror("エラー", f"メイン処理中にエラーが発生しました:\n{str(e)}")

if __name__ == "__main__":
    main()  # プログラムの実行
